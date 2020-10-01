# This file is part of Honeybee.
#
# Honeybee is free software: you can redistribute it and/or
# modify it under the terms of the GNU Affero General Public
# License as published by the Free Software Foundation, either
# version 3 of the License, or (at your option) any later
# version.
#
# Honeybee is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General
# Public License along with Honeybee.  If not, see
# <https://www.gnu.org/licenses/>.

import openpyxl


SURVEY_ORDER = (
    "type", "name", "label", "hint", "appearance", "constraint",
    "constraint_message", "calculation", "required",
    "required_message", "repeat_count")

CHOICES_ORDER = (
    "list_name", "value", "label", "filter")

SETTINGS_ORDER = (
    "form_title", "form_id", "version")


def column_key(order, value):
    try:
        return order.index(value)
    except ValueError:
        return len(order)


def get_column_names(rows):
    return set(name for row in rows
                    for name in row.keys())


def write_sheet(sheet, rows, column_key):
    column_names = sorted(
        get_column_names(rows), key=column_key)

    sheet.append(column_names)

    for row in rows:
        sheet.append(
            tuple(row.get(name) for name in column_names))


def write_xlsx(survey, filename):
    workbook = openpyxl.Workbook()

    survey_sheet = workbook.active
    survey_sheet.title = "survey"

    choices_sheet = workbook.create_sheet(title="choices")
    settings_sheet = workbook.create_sheet(title="settings")

    write_sheet(
        survey_sheet,
        survey.survey,
        lambda x: column_key(SURVEY_ORDER, x))
    write_sheet(
        choices_sheet,
        survey.choices,
        lambda x: column_key(CHOICES_ORDER, x))
    write_sheet(
        settings_sheet,
        survey.settings,
        lambda x: column_key(SETTINGS_ORDER, x))

    workbook.save(filename)
