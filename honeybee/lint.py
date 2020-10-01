# This file is part of SurveyKIT.
#
# SurveyKIT is free software: you can redistribute it and/or
# modify it under the terms of the GNU Affero General Public
# License as published by the Free Software Foundation, either
# version 3 of the License, or (at your option) any later
# version.
#
# SurveyKIT is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General
# Public License along with SurveyKIT.  If not, see
# <https://www.gnu.org/licenses/>.

import re

import argh
import openpyxl


def dict_rows(sheet):
    rows = sheet.iter_rows()
    header = tuple(cell.value for cell in next(rows))

    for row in rows:
        yield {key: str(cell.value or "")
               for key, cell in zip(header, row)}


def check_choices(choices):
    list_names = set()
    last_list_name = None

    for row in dict_rows(choices):
        list_name = row.get("list_name")

        if last_list_name is None:
            last_list_name = list_name
            continue

        if list_name in list_names and list_name != last_list_name:
            print(f"Warning: Choice list {list_name} is defined in multiple places.")


def choices_list_names(choices):
    return set(row.get("list_name") for row in dict_rows(choices))


def missing_survey_columns(survey):
    rows = survey.iter_rows()
    header = tuple(cell.value for cell in next(rows))

    return tuple(col for col in ("name", "type", "label")
                     if col not in header)


def defines_variable(row):
    t = row.get("type")

    return (t.strip() and t not in ("begin group", "end group",
                                    "begin repeat", "end repeat",
                                    "note"))


def defines_repeat(row):
    return row.get("type") == "begin repeat"


def defines_select(row):
    return (row.get("type").startswith("select_one ")
            or row.get("type").startswith("select_multiple "))


def defines_calculate(row):
    return row.get("type") in ("calculate", "calculate_here")


def extract_list_names(row):
    return tuple(list_name for list_name in row.get("type").split(" ")[1:]
                           if list_name and list_name != "or_other")


def non_required_question(row):
    return (defines_variable(row)
            and row.get("type") not in ("calculate",
                                        "calculate_here",
                                        "start",
                                        "end",
                                        "today",
                                        "deviceid",
                                        "subscriberid",
                                        "text audit")
            and row.get("required") != "yes")


def extract_variables(string):
    return tuple(re.findall(r"\${([^}]+)}", string))


def undefined_references(row, names):
    return tuple(variable for col in row
                          for variable in extract_variables(row.get(col))
                          if variable not in names)


def undefined_choice_lists(row, choices):
    return tuple(list_name for list_name in extract_list_names(row)
                           if list_name not in choices)


def missing_calculation(row):
    return row.get("calculation").strip() == ""


def check_survey(survey, choices):
    names = set()
    list_names = choices_list_names(choices)

    cols = missing_survey_columns(survey)
    if cols:
        print("Error: The survey worksheet is missing core columns ",
              ", ".join(cols),
              ".",
              sep="")
        return

    for row in dict_rows(survey):
        if defines_variable(row) or defines_repeat(row):
            if row.get("name") in names:
                print("Warning: %(name)s is defined more than once." % row)
            else:
                names.add(row.get("name"))

        if defines_select(row):
            lists = undefined_choice_lists(row, list_names)
            if lists:
                print("Error: %(name)s refers to undefined choice lists: " % row, end="")
                print("%s." % ", ".join(lists))

        if non_required_question(row):
            print("Warning: %(name)s is not required." % row)

        if defines_calculate(row) and missing_calculation(row):
            print("Error: %(name)s has empty calculation." % row)

        refs = undefined_references(row, names)
        if refs:
            print("Error: %(name)s has undefined references: " % row, end="")
            print("%s." % ", ".join(refs))


def main(filename):
    form = openpyxl.load_workbook(filename)

    choices = form.get_sheet_by_name("choices")
    survey = form.get_sheet_by_name("survey")

    check_choices(choices)
    check_survey(survey, choices)


def dispatch():
    argh.dispatch_command(main)


if __name__ == "__main__":
    dispatch()
